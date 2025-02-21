import os
import pandas as pd
import logging
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, numbers  # Importa números para formatos numéricos
from datetime import datetime

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('merge_csv.log'),
        logging.StreamHandler()
    ]
)


class SearsCsvMerger:
    def __init__(self):
        self.input_dir = 'CSVreporte'  # Carpeta donde se encuentran los archivos CSV
        self.concentrado_file = os.path.join('RESULTADOFINAL', 'Concentrado Sears.xlsx')
        self.backup_dir = os.path.join('RESULTADOFINAL', 'backups')
        self.report_file = os.path.join('RESULTADOFINAL', 'reporte_merge_csv.xlsx')
        
        # Mapeo de columnas del CSV a columnas del Excel (comenzando en AB)
        self.column_mapping = {
            'Pedido': 'M',            # Columna AB
            'Marketplace': 'N',       # Columna AC
            'Seller': 'O',            # Columna AD
            'Monto': 'P',             # Columna AE
            'Nombre_producto': 'Q',   # Columna AF
            'Precio': 'R',            # Columna AG
            'sku': 'S',               # Columna AH
            'Estatus_pedido': 'T',    # Columna AI
            'Estatus_partida': 'U',   # Columna AJ
            'Fecha_Pedido': 'V',      # Columna AK
            'IdFulfillment': 'W',     # Columna AL
            'NoGuia': 'X',            # Columna AM
            'Tipo_envio': 'Y'         # Columna AN
        }

    def create_backup(self):
        """Crea una copia de respaldo del archivo concentrado antes de modificarlo"""
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(self.backup_dir, f'Concentrado_Sears_backup_csv_{timestamp}.xlsx')
        
        if os.path.exists(self.concentrado_file):
            wb = load_workbook(self.concentrado_file)
            wb.save(backup_file)
            logging.info(f"Backup creado: {backup_file}")

    def get_cell_value(self, ws, row_idx, col_letter):
        """Obtiene el valor de una celda de manera segura"""
        cell = ws[f"{col_letter}{row_idx}"]
        return cell.value if cell else None

    def update_concentrado_cell(self, wb, sheet_name, row_idx, col_letter, value, csv_col):
        """Actualiza una celda específica preservando el formato solo si hay un cambio real"""
        ws = wb[sheet_name]
        
        # Obtener valor actual
        current_value = self.get_cell_value(ws, row_idx, col_letter)
        
        # Normalizar valores para comparación
        if isinstance(value, (int, float)) and pd.notna(value):
            current_float = float(current_value) if pd.notna(current_value) else None
            new_float = float(value)
            if current_float == new_float:
                return False, None  # No hay cambio real
        elif current_value == value:
            return False, None  # No hay cambio real
        
        # Si hay un cambio, actualizar la celda
        target_cell = ws[f"{col_letter}{row_idx}"]
        old_format = target_cell._style
        
        # Convertir el valor al tipo correcto
        if csv_col == 'Pedido':
            try:
                value = int(value)  # Forzar el pedido como número entero
            except ValueError:
                pass  # Si no se puede convertir, dejar como está
        
        # Manejar fechas
        if csv_col == 'Fecha_Pedido' and pd.notna(value):
            try:
                # Convertir la fecha a un objeto datetime
                parsed_date = pd.to_datetime(value)
                value = parsed_date  # Asignar el objeto datetime directamente
            except Exception as e:
                logging.warning(f"No se pudo formatear la fecha en columna {csv_col}: {str(e)}")
        
        target_cell.value = value
        
        # Reaplicar el estilo original
        target_cell._style = old_format

        # Aplicar formato numérico si es un número
        if isinstance(value, (int, float)):
            target_cell.number_format = numbers.FORMAT_NUMBER
        
        # Aplicar formato de fecha si es una fecha
        if csv_col == 'Fecha_Pedido' and pd.notna(value):
            try:
                target_cell.number_format = numbers.FORMAT_DATE_DDMMYYYY  # Formato de fecha en Excel
            except Exception as e:
                logging.warning(f"No se pudo aplicar el formato de fecha en columna {csv_col}: {str(e)}")
        
        return True, f"{csv_col}: {current_value} -> {value}"

    def merge_csv_data(self, csv_file):
        try:
            # Crear backup antes de comenzar
            self.create_backup()
            time.sleep(0.5)
            
            csv_filename = os.path.basename(csv_file)
            logging.info(f"Procesando archivo: {csv_filename}")
            
            # Leer el archivo CSV
            csv_df = pd.read_csv(csv_file, encoding='utf-8')
            csv_df['Pedido'] = csv_df['Pedido'].astype(str)
            
            # Leer el archivo concentrado
            logging.info("Leyendo archivo concentrado...")
            wb = load_workbook(self.concentrado_file)
            ws = wb.active
            concentrado_df = pd.read_excel(self.concentrado_file)
            concentrado_df['ORDEN SEARS '] = concentrado_df['ORDEN SEARS '].astype(str)
            
            # Contadores y listas
            updates = 0
            no_matches = 0
            matches = 0
            no_match_pedidos = []
            match_pedidos = []
            updated_pedidos = []  # Nueva lista para pedidos con cambios reales
            
            # Procesar cada fila del CSV
            for idx, row in csv_df.iterrows():
                time.sleep(0.05)
                pedido = str(row['Pedido'])
                
                # Buscar coincidencia
                mask = concentrado_df['ORDEN SEARS '] == pedido
                
                if mask.any():
                    excel_row_idx = mask.idxmax() + 2
                    changes = []
                    updates_in_row = 0
                    
                    # Verificar y actualizar cada campo
                    for csv_col, excel_col in self.column_mapping.items():
                        try:
                            value = row[csv_col]
                            if pd.notna(value):  # Solo procesar valores no nulos
                                updated, change_msg = self.update_concentrado_cell(
                                    wb, ws.title, excel_row_idx, excel_col, value, csv_col
                                )
                                if updated:
                                    updates_in_row += 1
                                    if change_msg:
                                        changes.append(change_msg)
                        except Exception as e:
                            logging.warning(f"Error en columna {csv_col}, pedido {pedido}: {str(e)}")
                    
                    matches += 1
                    match_pedidos.append(pedido)
                    
                    if updates_in_row > 0:
                        updates += 1
                        updated_pedidos.append(pedido)  # Agregar a lista de actualizados
                        logging.info(f"Pedido {pedido}: {updates_in_row} campos actualizados")
                        if changes:
                            logging.info("Cambios: " + ", ".join(changes))
                    
                else:
                    no_matches += 1
                    no_match_pedidos.append(pedido)
                    logging.warning(f"No se encontró coincidencia para el pedido: {pedido}")
            
            # Guardar archivo actualizado
            logging.info("Guardando archivo actualizado...")
            wb.save(self.concentrado_file)
            
            # Determinar el número máximo de filas necesarias
            max_rows = max(len(match_pedidos), len(updated_pedidos), len(no_match_pedidos))
            
            # Crear el reporte con tres columnas y un solo encabezado
            report_data = {
                'Registros encontrados': match_pedidos + [''] * (max_rows - len(match_pedidos)),
                'Registros actualizados': updated_pedidos + [''] * (max_rows - len(updated_pedidos)),
                'Registros sin coincidencia': no_match_pedidos + [''] * (max_rows - len(no_match_pedidos))
            }
            report_df = pd.DataFrame(report_data)
            
            # Guardar el reporte con encabezados en negritas y formato numérico en las celdas
            with pd.ExcelWriter(self.report_file, engine='openpyxl') as writer:
                report_df.to_excel(writer, index=False, sheet_name='Reporte')
                workbook = writer.book
                worksheet = writer.sheets['Reporte']
                
                # Aplicar formato en negritas al encabezado
                bold_font = Font(bold=True)
                for col, header in enumerate(report_df.columns, 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.value = header
                    cell.font = bold_font
                
                # Aplicar formato numérico a las columnas específicas
                numeric_columns = ['Registros encontrados', 'Registros actualizados', 'Registros sin coincidencia']
                for col_idx, col_name in enumerate(report_df.columns, 1):
                    if col_name in numeric_columns:  # Solo aplica formato a las columnas numéricas
                        for row_idx in range(2, len(report_df) + 2):  # Itera sobre las filas de datos (fila 2 en adelante)
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            try:
                                # Intenta convertir el valor a número y aplica formato
                                cell.value = float(cell.value) if cell.value else None
                                cell.number_format = numbers.FORMAT_NUMBER  # Formato de número entero
                            except ValueError:
                                # Si no se puede convertir, deja la celda como está
                                pass
            
            # Resumen en logs
            logging.info(f"""
            Resumen del proceso de merge CSV:
            Archivo: {csv_filename}
            - Total de registros: {len(csv_df)}
            - Registros encontrados: {matches}
            - Registros actualizados: {updates}
            - Registros sin coincidencia: {no_matches}
            - Pedidos encontrados: {', '.join(match_pedidos)}
            - Pedidos no encontrados: {', '.join(no_match_pedidos)}
            
            El reporte detallado se ha guardado en: {self.report_file}
            """)
            
        except Exception as e:
            logging.error(f"Error durante el proceso de merge CSV: {str(e)}")
            raise

    def process_all_csvs(self):
        """Procesa todos los CSVs en la carpeta CSVreporte"""
        if not os.path.exists(self.input_dir):
            os.makedirs(self.input_dir)
            logging.info(f"Carpeta {self.input_dir} creada")
            return
        
        csv_files = [f for f in os.listdir(self.input_dir) if f.endswith('.csv')]
        if not csv_files:
            logging.info("No se encontraron archivos CSV para procesar")
            return
        
        for csv_file in sorted(csv_files):
            file_path = os.path.join(self.input_dir, csv_file)
            self.merge_csv_data(file_path)
            time.sleep(1)  # Delay entre archivos

if __name__ == "__main__":
    merger = SearsCsvMerger()
    merger.process_all_csvs()