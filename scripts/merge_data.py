import os
import pandas as pd
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, numbers

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join('logs', 'merge.log')), # Changed log file path
        logging.StreamHandler()
    ]
)

class SearsMerger:
    def __init__(self):
        self.output_file = os.path.join('EXCELPDFSEARS', 'sears_extractions.xlsx')
        self.concentrado_file = os.path.join('RESULTADOFINAL', 'Concentrado Sears.xlsx')
        self.backup_dir = os.path.join('RESULTADOFINAL', 'backups')
        self.report_file = os.path.join('RESULTADOFINAL', 'reporte_merge.xlsx')

    def create_backup(self):
        """Crea una copia de respaldo del archivo concentrado antes de modificarlo"""
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(self.backup_dir, f'Concentrado_Sears_backup_{timestamp}.xlsx')
        
        if os.path.exists(self.concentrado_file):
            wb = load_workbook(self.concentrado_file)
            wb.save(backup_file)
            logging.info(f"Backup creado: {backup_file}")

    def process_duplicates(self, extractions_df):
        """Procesa los pedidos duplicados, sumando sus totales"""
        # Identificar duplicados
        duplicados = extractions_df[extractions_df.duplicated(['Numero_Pedido'], keep=False)]
        pedidos_duplicados = duplicados['Numero_Pedido'].unique()
        
        # Diccionario para almacenar los resultados procesados
        processed_data = {}
        
        for pedido in pedidos_duplicados:
            registros = extractions_df[extractions_df['Numero_Pedido'] == pedido]
            
            # Sumar los totales
            total_sumado = registros['Total'].sum()
            
            # Tomar los datos del primer registro
            primer_registro = registros.iloc[0].to_dict()
            
            # Actualizar el total y agregar nota sobre la suma
            primer_registro['Total'] = total_sumado
            primer_registro['documentos_sumados'] = ', '.join(registros['Numero_Documento'].astype(str))
            
            # Guardar en el diccionario de procesados
            processed_data[pedido] = primer_registro
            
            logging.info(f"""
            Pedido duplicado procesado: {pedido}
            Documentos sumados: {primer_registro['documentos_sumados']}
            Total sumado: {total_sumado}
            """)
        
        return processed_data, pedidos_duplicados

    def merge_data(self):
        try:
            # Crear backup antes de comenzar
            self.create_backup()
            
            # Leer el archivo de extracciones
            logging.info("Leyendo archivo de extracciones...")
            extractions_df = pd.read_excel(self.output_file)
            
            # Convertir columnas de fecha a datetime
            date_columns = ['Fecha_Pedido', 'Fecha_Vencimiento']
            for col in date_columns:
                if col in extractions_df.columns:
                    extractions_df[col] = pd.to_datetime(extractions_df[col], errors='coerce')  # Convertir a datetime
            
            # Leer el archivo concentrado
            logging.info("Leyendo archivo concentrado...")
            concentrado_df = pd.read_excel(self.concentrado_file)
            
            # Asegurar tipos de datos correctos
            extractions_df['Numero_Pedido'] = extractions_df['Numero_Pedido'].astype(str)
            concentrado_df['ORDEN SEARS '] = concentrado_df['ORDEN SEARS '].astype(str)
            
            # Procesar duplicados
            processed_duplicates, pedidos_duplicados = self.process_duplicates(extractions_df)
            
            # Contador para seguimiento
            updates = 0
            no_matches = 0
            
            # Cargar el archivo existente con openpyxl
            wb = load_workbook(self.concentrado_file)
            ws = wb.active
            
            # Obtener el mapeo de columnas por nombre
            column_mapping = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}  # Mapeo de nombres a índices
            
            # Iterar sobre las filas del archivo de extracciones
            for idx, row in extractions_df.iterrows():
                pedido = row['Numero_Pedido']
                
                # Si es un duplicado y ya lo procesamos, saltarlo
                if pedido in pedidos_duplicados and pedido not in processed_duplicates:
                    continue
                
                # Buscar coincidencia en el archivo concentrado
                mask = concentrado_df['ORDEN SEARS '] == pedido
                if mask.any():
                    # Encontrar la fila correspondiente
                    row_idx = mask.idxmax() + 2  # +2 porque Excel usa 1-based indexing y tiene encabezado
                    
                    if pedido in processed_duplicates:
                        datos = processed_duplicates[pedido]
                        # Actualizar campos usando el mapeo de columnas
                        ws.cell(row=row_idx, column=column_mapping['Total']).value = datos['Total']
                        ws.cell(row=row_idx, column=column_mapping['OBSERVACIONES ']).value = (
                            f"SUMA DE PRODUCTOS - Documentos: {datos['documentos_sumados']}"
                        )
                        
                        # Formatear fechas si existen
                        if 'Fecha_Pedido' in datos and pd.notna(datos['Fecha_Pedido']):
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Pedido']).value = datos['Fecha_Pedido']
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Pedido']).number_format = "dd/mm/yyyy"
                        if 'Fecha_Vencimiento' in datos and pd.notna(datos['Fecha_Vencimiento']):
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Vencimiento']).value = datos['Fecha_Vencimiento']
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Vencimiento']).number_format = "dd/mm/yyyy"
                        
                        updates += 1
                        logging.info(f"""
                        Actualizado pedido duplicado: {pedido}
                        Total sumado: {datos['Total']}
                        Documentos: {datos['documentos_sumados']}
                        """)
                    else:
                        # Actualizar campos usando el mapeo de columnas
                        ws.cell(row=row_idx, column=column_mapping['Total']).value = row['Total']
                        if pd.notna(row['Fecha_Pedido']):
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Pedido']).value = row['Fecha_Pedido']
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Pedido']).number_format = "dd/mm/yyyy"
                        if pd.notna(row['Fecha_Vencimiento']):
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Vencimiento']).value = row['Fecha_Vencimiento']
                            ws.cell(row=row_idx, column=column_mapping['Fecha_Vencimiento']).number_format = "dd/mm/yyyy"
                        ws.cell(row=row_idx, column=column_mapping['Numero_Documento']).value = int(row['Numero_Documento']) if pd.notna(row['Numero_Documento']) else None
                        ws.cell(row=row_idx, column=column_mapping['Tipo_Docto']).value = row['Tipo_Docto']
                        ws.cell(row=row_idx, column=column_mapping['Descripcion']).value = row['Descripcion']
                        ws.cell(row=row_idx, column=column_mapping['Cheque']).value = row['Cheque']
                        ws.cell(row=row_idx, column=column_mapping['Proveedor']).value = row['Proveedor']
                        
                        updates += 1
                else:
                    no_matches += 1
                    logging.warning(f"No se encontró coincidencia para el pedido: {pedido}")
            
            # Guardar el archivo actualizado
            logging.info("Guardando archivo actualizado...")
            wb.save(self.concentrado_file)
            
            # Resumen del proceso
            logging.info(f"""
            Resumen del proceso de merge:
            - Total de registros procesados: {len(extractions_df)}
            - Registros actualizados: {updates}
            - Registros sin coincidencia: {no_matches}
            """)
            logging.info("Proceso de merge completado exitosamente")
        except Exception as e:
            logging.error(f"Error durante el proceso de merge: {str(e)}")
            raise

if __name__ == "__main__":
    merger = SearsMerger()
    merger.merge_data()
